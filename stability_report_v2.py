"""
Stability Report Generator v2
==============================
Reads raw Raven output files directly. No manual summary step needed.

File naming convention (REQUIRED):
    {Instrument}_{Run}_{Lot}_{Timepoint}.xlsx
    {Instrument}_{Run}_{Lot}_{Timepoint}_Rerun1.xlsx
    {Instrument}_{Run}_{Lot}_{Timepoint}_Rerun2.xlsx

Folder structure (REQUIRED):
    StudyFolder/          e.g. HE4_Stab11/
        TP1/
            A4_Run1_Lot1_TP1.xlsx
            A4_Run1_Lot2_TP1.xlsx
            A4_Run1_Validity_TP1.xlsx
            ... (18 files per timepoint minimum)
        TP2/
            ...
"""

import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# SPEC LIMITS — loaded from config.xlsx
# These are fallback defaults used only if no
# config file is found (HE4 Stab 11 values).
# ─────────────────────────────────────────────

_DEFAULT_INDIV_REP_SPECS = {
    "Panel1":  {"LSL": 34.7,  "USL": 54.9},
    "Panel2":  {"LSL": 122.0, "USL": 192.8},
    "Panel3":  {"LSL": 355.9, "USL": 562.5},
    "Panel4":  {"LSL": 877.8, "USL": 1387.4},
    "RefCon1": {"LSL": 28.6,  "USL": 45.2},
    "RefCon2": {"LSL": 115.9, "USL": 183.1},
    "RefCon3": {"LSL": 504.1, "USL": 796.9},
}

_DEFAULT_MEAN_SPECS = {
    "Panel1":  {"LSL": 41.4,   "USL": 48.2},
    "Panel2":  {"LSL": 145.6,  "USL": 169.2},
    "Panel3":  {"LSL": 424.8,  "USL": 493.6},
    "Panel4":  {"LSL": 1047.7, "USL": 1217.5},
    "RefCon1": {"LSL": 34.1,   "USL": 39.7},
    "RefCon2": {"LSL": 138.3,  "USL": 160.7},
    "RefCon3": {"LSL": 601.7,  "USL": 699.3},
}

_DEFAULT_VALIDITY_SPECS = {
    "RefCon1": {"LSL": 28.6,  "USL": 45.2},
    "RefCon2": {"LSL": 115.9, "USL": 183.1},
    "RefCon3": {"LSL": 504.1, "USL": 796.9},
}

_DEFAULT_CV_LIMIT  = 7.5
_DEFAULT_MIN_REPS  = 10

# Active specs — replaced by load_config() when config.xlsx is present
INDIV_REP_SPECS = _DEFAULT_INDIV_REP_SPECS.copy()
MEAN_SPECS      = _DEFAULT_MEAN_SPECS.copy()
VALIDITY_SPECS  = _DEFAULT_VALIDITY_SPECS.copy()
CV_LIMIT        = _DEFAULT_CV_LIMIT
MIN_REPS        = _DEFAULT_MIN_REPS
CONC_DECIMALS   = 1   # decimal places for concentration values (configurable per assay)

# Active samples — derived from config or defaults
SAMPLES_OF_INTEREST = list(_DEFAULT_INDIV_REP_SPECS.keys())
SAMPLE_ORDER        = SAMPLES_OF_INTEREST[:]

# Per-timepoint spec lookup:
# {(sample, lot): [(effective_from_tp_num, indiv_spec, mean_spec, validity_spec,
#                   within_run_cv, between_run_cv), ...]}
# sorted by tp_num ascending so we can find the applicable spec for any timepoint
_TIMEPOINT_SPECS = {}

# Per-sample CV limits — {(sample, lot): within_run_cv, between_run_cv}
# Falls back to CV_LIMIT global if not specified in config
_SAMPLE_CV_LIMITS = {}




def format_run_dates(dates):
    """
    Format Raven date strings into a readable display.
    Handles single date or range if files span multiple days.
    Raven format: '2026-05-21T12.25'
    """
    from datetime import datetime
    parsed = set()
    for d in dates:
        if not d:
            continue
        try:
            # Raven uses 2026-05-21T12.25 format
            date_part = str(d).split("T")[0]
            parsed.add(datetime.strptime(date_part, "%Y-%m-%d").date())
        except ValueError:
            pass

    if not parsed:
        return str(sorted(dates)[0]) if dates else "N/A"

    sorted_dates = sorted(parsed)
    fmt = lambda d: d.strftime("%d %b %Y")
    if len(sorted_dates) == 1:
        return fmt(sorted_dates[0])
    return f"{fmt(sorted_dates[0])} \u2013 {fmt(sorted_dates[-1])}"


# ─────────────────────────────────────────────
# CONFIG LOADER
# ─────────────────────────────────────────────

def load_config(study_folder):
    """
    Read config.xlsx from the study folder and populate global spec variables.
    Returns (assay_name, study_number, config_loaded).
    If no config found, returns defaults and warns.
    """
    global INDIV_REP_SPECS, MEAN_SPECS, VALIDITY_SPECS
    global CV_LIMIT, MIN_REPS, CONC_DECIMALS
    global SAMPLES_OF_INTEREST, SAMPLE_ORDER
    global _TIMEPOINT_SPECS, _SAMPLE_CV_LIMITS

    config_path = os.path.join(study_folder, "config.xlsx")

    if not os.path.exists(config_path):
        print(f"  ⚠  No config.xlsx found in {study_folder} — using built-in HE4 defaults.")
        return "HE4", "Stab 11", False

    print(f"  📋 Loading config from {config_path}")

    # ── Study Info sheet ──
    try:
        info_df = pd.read_excel(config_path, sheet_name="Study Info",
                                header=2, usecols=[1, 2])
        info_df.columns = ["Field", "Value"]
        info_df = info_df.dropna(subset=["Field"])
        info    = dict(zip(info_df["Field"].str.strip(),
                           info_df["Value"]))

        assay_name   = str(info.get("Assay Name",   "Unknown")).strip()
        study_number = str(info.get("Study Number", "Stab")).strip()
        cv_lim       = float(info.get("CV Limit (%)",    7.5))
        min_reps     = int(float(info.get("Min Valid Reps", 10)))
        conc_dec     = int(float(info.get("Concentration Decimal Places", 1)))
    except Exception as e:
        print(f"  ⚠  Could not read Study Info sheet: {e} — using defaults.")
        assay_name, study_number = "Unknown", "Stab"
        cv_lim, min_reps, conc_dec = 7.5, 10, 1

    CV_LIMIT      = cv_lim
    MIN_REPS      = min_reps
    CONC_DECIMALS = conc_dec

    # ── Specs sheet ──
    try:
        specs_df = pd.read_excel(config_path, sheet_name="Specs",
                                 header=3)   # row 4 is the column header row
        # Normalise column names
        specs_df.columns = [str(c).strip() for c in specs_df.columns]

        # Map flexible column names to standard keys
        col_map = {
            "Sample":            "Sample",
            "Lot":               "Lot",
            "Effective\nFrom":  "EffectiveFrom",
            "Effective From":    "EffectiveFrom",
            "Indiv Rep\nLSL":  "IndivLSL",
            "Indiv Rep LSL":     "IndivLSL",
            "Indiv Rep\nUSL":  "IndivUSL",
            "Indiv Rep USL":     "IndivUSL",
            "Mean\nLSL":       "MeanLSL",
            "Mean LSL":          "MeanLSL",
            "Mean\nUSL":       "MeanUSL",
            "Mean USL":          "MeanUSL",
            "Validity\nLSL":   "ValidityLSL",
            "Validity LSL":      "ValidityLSL",
            "Validity\nUSL":   "ValidityUSL",
            "Validity USL":      "ValidityUSL",
            "Within-Run\nCV (%)": "WithinRunCV",
            "Within-Run CV (%)":   "WithinRunCV",
            "Between-Run\nCV (%)": "BetweenRunCV",
            "Between-Run CV (%)":   "BetweenRunCV",
        }
        specs_df.rename(columns=col_map, inplace=True)
        specs_df = specs_df.dropna(subset=["Sample", "Lot"])

        # Build _TIMEPOINT_SPECS lookup
        _TIMEPOINT_SPECS = {}
        for _, row in specs_df.iterrows():
            sample = str(row["Sample"]).strip()
            lot    = str(row["Lot"]).strip()
            ef_raw = str(row.get("EffectiveFrom", "TP1")).strip()
            ef_num = int(re.sub(r"[^0-9]", "", ef_raw) or "1")

            indiv_spec = None
            if pd.notna(row.get("IndivLSL")) and pd.notna(row.get("IndivUSL")):
                indiv_spec = {"LSL": float(row["IndivLSL"]),
                              "USL": float(row["IndivUSL"])}

            mean_spec = None
            if pd.notna(row.get("MeanLSL")) and pd.notna(row.get("MeanUSL")):
                mean_spec = {"LSL": float(row["MeanLSL"]),
                             "USL": float(row["MeanUSL"])}

            validity_spec = None
            if pd.notna(row.get("ValidityLSL")) and pd.notna(row.get("ValidityUSL")):
                validity_spec = {"LSL": float(row["ValidityLSL"]),
                                 "USL": float(row["ValidityUSL"])}

            key = (sample, lot)
            if key not in _TIMEPOINT_SPECS:
                _TIMEPOINT_SPECS[key] = []
            _TIMEPOINT_SPECS[key].append((ef_num, indiv_spec, mean_spec, validity_spec))

        # Sort each entry by effective_from ascending
        for key in _TIMEPOINT_SPECS:
            _TIMEPOINT_SPECS[key].sort(key=lambda x: x[0])

        # Build per-sample CV limits lookup
        _SAMPLE_CV_LIMITS = {}
        for _, row in specs_df.iterrows():
            sample = str(row["Sample"]).strip()
            lot    = str(row["Lot"]).strip()
            wrcv   = float(row["WithinRunCV"])  if pd.notna(row.get("WithinRunCV"))  else CV_LIMIT
            brcv   = float(row["BetweenRunCV"]) if pd.notna(row.get("BetweenRunCV")) else CV_LIMIT
            key = (sample, lot)
            if key not in _SAMPLE_CV_LIMITS:
                _SAMPLE_CV_LIMITS[key] = (wrcv, brcv)

        # Build flat spec dicts for TP1 (used as defaults throughout)
        # These represent the most common/initial spec for each sample
        new_indiv    = {}
        new_mean     = {}
        new_validity = {}
        all_samples  = set()

        for (sample, lot), entries in _TIMEPOINT_SPECS.items():
            all_samples.add(sample)
            # Use TP1 (first) entry for the flat dicts
            _, indiv, mean, validity = entries[0]
            if indiv and sample not in new_indiv:
                new_indiv[sample] = indiv
            if mean and sample not in new_mean:
                new_mean[sample] = mean
            if validity and sample not in new_validity:
                new_validity[sample] = validity

        INDIV_REP_SPECS = new_indiv
        MEAN_SPECS      = new_mean
        VALIDITY_SPECS  = new_validity

        # Update sample order from config (panels first, then refcons)
        panels  = sorted([s for s in all_samples if s.startswith("Panel")],
                         key=lambda x: int(re.sub(r"[^0-9]", "", x) or "0"))
        refcons = sorted([s for s in all_samples if s.startswith("RefCon")],
                         key=lambda x: int(re.sub(r"[^0-9]", "", x) or "0"))
        others  = sorted([s for s in all_samples
                          if not s.startswith("Panel") and
                          not s.startswith("RefCon")])
        SAMPLES_OF_INTEREST = panels + refcons + others
        SAMPLE_ORDER        = SAMPLES_OF_INTEREST[:]

        print(f"  ✅ Config loaded: {assay_name} {study_number}")
        print(f"     Default CV limit: {CV_LIMIT}%  |  Min reps: {MIN_REPS}  |  Conc decimals: {CONC_DECIMALS}")
        print(f"     Per-sample CV limits loaded: {len(_SAMPLE_CV_LIMITS)} entries")
        print(f"     Samples: {SAMPLES_OF_INTEREST}")
        print(f"     Spec rows: {len(specs_df)}")

    except Exception as e:
        print(f"  ⚠  Could not read Specs sheet: {e} — using defaults.")

    return assay_name, study_number, True


def get_cv_limits(sample, lot="Lot1"):
    """
    Return (within_run_cv_pct, between_run_cv_pct) for a given sample/lot.
    Falls back to global CV_LIMIT if not specified per-sample in config.
    """
    key = (sample, lot)
    if key in _SAMPLE_CV_LIMITS:
        return _SAMPLE_CV_LIMITS[key]
    # Try any lot for this sample
    for k, v in _SAMPLE_CV_LIMITS.items():
        if k[0] == sample:
            return v
    return (CV_LIMIT, CV_LIMIT)


def _get_br_cv_pf(cv_val, sample, lot="Lot1"):
    """Return PASS/FAIL for a between-run CV value against per-sample between-run limit."""
    if cv_val is None:
        return None
    _, between_cv = get_cv_limits(sample, lot)
    # cv_val is stored as decimal (e.g. 0.0073), multiply by 100 to get percent
    return "PASS" if cv_val * 100 <= between_cv else "FAIL"


def get_specs_for(sample, lot, timepoint):
    """
    Return (indiv_spec, mean_spec, validity_spec) for a given
    sample/lot/timepoint, respecting mid-study lot changes.
    Falls back to the flat dicts if no timepoint-specific entry exists.
    """
    tp_num = int(re.sub(r"[^0-9]", "", str(timepoint)) or "1")
    key    = (sample, lot)

    if key in _TIMEPOINT_SPECS:
        # Find the most recent entry whose effective_from <= tp_num
        applicable = None
        for ef_num, indiv, mean, validity in _TIMEPOINT_SPECS[key]:
            if ef_num <= tp_num:
                applicable = (indiv, mean, validity)
            else:
                break
        if applicable:
            return applicable

    # Fallback to flat dicts
    return (INDIV_REP_SPECS.get(sample),
            MEAN_SPECS.get(sample),
            VALIDITY_SPECS.get(sample) if sample.startswith("RefCon") else None)


# ─────────────────────────────────────────────
# FILENAME PARSER
# ─────────────────────────────────────────────

def parse_filename(fname):
    """
    Parse a Raven output filename into its components.

    Expected formats:
        A4_Run1_Lot1_TP1.xlsx
        A4_Run1_Lot2_TP1_Rerun1.xlsx
        A4_Run1_Validity_TP1_Rerun2.xlsx

    Returns dict with keys:
        instrument, run_num, lot, timepoint, rerun_num
        rerun_num is 0 for originals, 1/2 for reruns
    """
    stem = os.path.splitext(fname)[0]  # strip .xlsx

    # Check for rerun suffix
    rerun_match = re.search(r'_Rerun(\d+)$', stem, re.IGNORECASE)
    rerun_num = int(rerun_match.group(1)) if rerun_match else 0
    if rerun_match:
        stem = stem[:rerun_match.start()]  # strip rerun suffix

    # Now parse: Instrument_Run#_Lot_TP#
    pattern = r'^([A-Za-z]\d+)_Run(\d+)_(Lot\d+|Validity)_(TP\d+)$'
    m = re.match(pattern, stem, re.IGNORECASE)
    if not m:
        return None

    instrument_code = m.group(1).upper()   # e.g. A4
    run_num         = int(m.group(2))       # e.g. 1
    lot             = m.group(3)            # e.g. Lot1, Lot2, Validity
    # Normalise capitalisation
    if lot.lower() == "validity":
        lot = "Validity"
    else:
        lot = lot[0].upper() + lot[1:].lower()  # Lot1, Lot2
    timepoint = m.group(4).upper()          # e.g. TP1

    # Instrument display name: A4 → AIM 4
    instr_num = re.sub(r'[^0-9]', '', instrument_code)
    instrument = f"AIM {instr_num}"

    return {
        "instrument":      instrument,
        "instrument_code": instrument_code,
        "run_num":         run_num,
        "lot":             lot,
        "lot_type":        "Validity" if lot == "Validity" else "Feasibility",
        "timepoint":       timepoint,
        "rerun_num":       rerun_num,
        "filename":        fname,
    }


# ─────────────────────────────────────────────
# RAVEN FILE READER
# ─────────────────────────────────────────────

def read_raven_file(path):
    """
    Read one raw Raven output file.
    Returns (meta dict, replicates DataFrame).
    meta keys: study_name_raven, run_date
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Analysis"]
    rows = list(ws.iter_rows(max_row=30, values_only=True))

    # Pull metadata from header rows
    study_name_raven = None
    run_date         = None
    for row in rows:
        if row[0] == "Study Name":
            study_name_raven = str(row[1]) if row[1] else None
        if row[0] == "Creation D/T":
            run_date = str(row[1]) if row[1] else None

    # Find data header row (has System, TDEF, SampleID, Rep, RLU, Result, Flag)
    header_row_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "System" in vals and "SampleID" in vals and "Rep" in vals:
            header_row_idx = i + 1  # 1-based for pandas
            break

    if header_row_idx is None:
        return None, pd.DataFrame()

    df = pd.read_excel(path, sheet_name="Analysis", header=header_row_idx - 1)
    df.columns = df.columns.str.strip()

    keep = ["System", "TDEF", "Lot", "SampleID", "Date",
            "RunID", "Rep", "RLU", "Result", "Flag"]
    df = df[[c for c in keep if c in df.columns]].copy()

    # Filter to samples of interest only
    # Warn about unexpected sample names that look like Panel/RefCon typos
    # Silently skip known non-sample rows (calibrators, scalars, etc.)
    EXPECTED_NON_SAMPLES = {"Scal", "Cal", "Scale", "Calibrator", "Blank"}
    all_samples = df["SampleID"].dropna().unique()
    unknown = [
        s for s in all_samples
        if s not in SAMPLES_OF_INTEREST
        and not any(s.startswith(prefix) for prefix in EXPECTED_NON_SAMPLES)
    ]
    if unknown:
        recognised = [s for s in all_samples if s in SAMPLES_OF_INTEREST]
        print(f"  ⚠  Unrecognised sample name(s) in {os.path.basename(path)}: "
              f"{unknown}")
        print(f"     Recognised in same file: {recognised}")
        print(f"     These rows will be excluded from the report. This is likely a "
              f"data entry issue in Raven that must be resolved at the source.")
    df = df[df["SampleID"].isin(SAMPLES_OF_INTEREST)].copy()

    df["BelowCurve"] = df["Result"].astype(str).str.contains("LOGIT", case=False)
    df["Result"]     = pd.to_numeric(df["Result"], errors="coerce")
    df["RLU"]        = pd.to_numeric(df["RLU"],    errors="coerce")
    df["Flag"]       = pd.to_numeric(df["Flag"],   errors="coerce").fillna(0).astype(int)

    meta = {
        "study_name_raven": study_name_raven,
        "run_date":         run_date,
    }

    return meta, df


# ─────────────────────────────────────────────
# LOAD ALL FILES FOR ONE TIMEPOINT
# ─────────────────────────────────────────────

def load_timepoint(tp_folder, timepoint, study_name):
    """
    Load all Raven files from one TP folder.
    Returns (summary_df, reps_df) for this timepoint.

    Rerun logic:
    - Original file: rerun_num = 0
    - Rerun files:   rerun_num = 1, 2, ...
    - For each (instrument, run_num, lot, sample):
        * If a rerun exists, it supersedes the original for that sample
        * Original is kept in records but flagged as superseded
        * For best-2-of-3: use whichever 2 of the 3 versions pass
    """
    all_files = [f for f in os.listdir(tp_folder)
                 if f.endswith(".xlsx") and not f.startswith("~")]

    # Parse and group files
    file_records = []
    for fname in sorted(all_files):
        parsed = parse_filename(fname)
        if parsed is None:
            print(f"  ⚠ Skipping unrecognised filename: {fname}")
            continue
        if parsed["timepoint"] != timepoint:
            print(f"  ⚠ Timepoint mismatch in {fname} — expected {timepoint}")
            continue
        parsed["path"] = os.path.join(tp_folder, fname)
        file_records.append(parsed)

    if not file_records:
        return pd.DataFrame(), pd.DataFrame()

    # Group by (instrument, run_num, lot) — each group may have original + reruns
    from collections import defaultdict
    run_groups = defaultdict(list)
    for rec in file_records:
        key = (rec["instrument"], rec["run_num"], rec["lot"])
        run_groups[key].append(rec)

    # Sort each group by rerun_num so original (0) comes first
    for key in run_groups:
        run_groups[key].sort(key=lambda x: x["rerun_num"])

    summary_rows = []
    rep_rows     = []

    for (instrument, run_num, lot), versions in run_groups.items():
        lot_type = "Validity" if lot == "Validity" else "Feasibility"

        # Read all versions
        version_data = {}  # rerun_num → {sample → rep rows}
        run_dates = set()
        study_name_raven = None

        for rec in versions:
            try:
                meta, df = read_raven_file(rec["path"])
            except Exception as e:
                fname = os.path.basename(rec['path'])
                # Detect file-locked-by-Excel specifically
                err_str = str(e).lower()
                if any(x in err_str for x in
                       ['permission', 'access', 'locked', 'being used',
                        'cannot access', 'winerror 32', 'winerror 13']):
                    print(f"  ⚠  Cannot read {fname} — file appears to be open "
                          f"in Excel. Close it and rerun the generator.")
                else:
                    print(f"  ⚠  Could not read {fname}: "
                          f"{type(e).__name__} — file will be skipped.")
                continue
            if df.empty:
                continue
            if meta.get("run_date"):
                run_dates.add(meta.get("run_date"))
            if study_name_raven is None:
                study_name_raven = meta.get("study_name_raven")

            version_data[rec["rerun_num"]] = df

        run_date = format_run_dates(run_dates)
        if not version_data:
            continue

        # Determine which version to use per sample
        # Priority: highest rerun_num that has data for that sample
        # For best-2-of-3 (rerun_num 0,1,2): pick 2 passing versions
        all_versions_sorted = sorted(version_data.keys())
        original_df = version_data.get(0, pd.DataFrame())
        all_samples = set()
        for df in version_data.values():
            all_samples.update(df["SampleID"].unique())
        all_samples = [s for s in SAMPLES_OF_INTEREST if s in all_samples]

        for sample in all_samples:
            # Collect this sample's data from each version
            sample_versions = {}
            for ver_num, df in version_data.items():
                sdf = df[df["SampleID"] == sample].copy()
                if not sdf.empty:
                    sample_versions[ver_num] = sdf

            if not sample_versions:
                continue

            ver_nums = sorted(sample_versions.keys())
            is_rerun = len(ver_nums) > 1

            # Determine which version is "accepted" for calculations
            if len(ver_nums) == 1:
                # Only original, no rerun
                accepted_ver = ver_nums[0]
                superseded_vers = []
            elif len(ver_nums) == 2:
                # Original + one rerun — rerun supersedes original
                accepted_ver  = ver_nums[1]
                superseded_vers = [ver_nums[0]]
            else:
                # 3 versions — best 2 of 3
                # Evaluate each version
                def passes(sdf):
                    valid = sdf[(sdf["Flag"] == 1) & (~sdf["BelowCurve"])]
                    if len(valid) < 1:
                        return False
                    mean = valid["Result"].mean()
                    spec = (VALIDITY_SPECS if lot_type == "Validity"
                            else MEAN_SPECS).get(sample, {})
                    if not spec:
                        return True
                    return spec["LSL"] <= mean <= spec["USL"]

                pass_status = {v: passes(sample_versions[v]) for v in ver_nums}
                passing = [v for v in ver_nums if pass_status[v]]
                failing = [v for v in ver_nums if not pass_status[v]]

                if len(passing) >= 2:
                    # Use the two most recent passing versions
                    accepted_vers  = sorted(passing)[-2:]
                    superseded_vers = [v for v in ver_nums
                                       if v not in accepted_vers]
                    accepted_ver   = accepted_vers[-1]  # primary for summary
                else:
                    # Can't get 2 passing — use latest anyway, flag as fail
                    accepted_ver   = ver_nums[-1]
                    superseded_vers = ver_nums[:-1]

            # Build run_id display string
            rerun_label = ""
            if accepted_ver > 0:
                rerun_label = f" (Rerun {accepted_ver})"

            run_id = f"Run{run_num}{rerun_label}"

            # Stats for the accepted version
            accepted_df = sample_versions[accepted_ver]
            valid = accepted_df[(accepted_df["Flag"] == 1) &
                                (~accepted_df["BelowCurve"])]
            total_reps = len(accepted_df)
            valid_reps = len(valid)

            if valid_reps >= 1:
                mean_rlu  = accepted_df[accepted_df["Flag"]==1]["RLU"].mean()
                mean_dose = valid["Result"].mean()
            else:
                mean_rlu = mean_dose = None

            if valid_reps >= 2:
                std_rlu  = accepted_df[accepted_df["Flag"]==1]["RLU"].std()
                std_dose = valid["Result"].std()
                rlu_cv   = (std_rlu  / mean_rlu  * 100) if mean_rlu  else None
                dose_cv  = (std_dose / mean_dose * 100) if mean_dose else None
            else:
                rlu_cv = dose_cv = None

            summary_rows.append({
                "Study":        study_name,
                "Timepoint":    timepoint,
                "Instrument":   instrument,
                "RunNum":       run_num,
                "RunID":        run_id,
                "Lot":          lot,
                "LotType":      lot_type,
                "Sample":       sample,
                "RunDate":      run_date,
                "Reps":         f"{valid_reps} / {total_reps}",
                "RLU_Avg":      round(mean_rlu,  0) if mean_rlu  is not None else None,
                "Dose_Avg":     round_conc(mean_dose) if mean_dose is not None else None,
                "RLU_CV_pct":   round(rlu_cv,   4) if rlu_cv    is not None else None,
                "Dose_CV_pct":  round(dose_cv,  4) if dose_cv   is not None else None,
                "ValidReps":    valid_reps,
                "TotalReps":    total_reps,
                "MeanDose":     mean_dose,
                "IsRerun":      is_rerun,
                "AcceptedVer":  accepted_ver,
            })

            # Store individual reps for accepted version
            for _, rep_row in valid.iterrows():
                rep_rows.append({
                    "Study":      study_name,
                    "Timepoint":  timepoint,
                    "Instrument": instrument,
                    "RunNum":     run_num,
                    "RunID":      run_id,
                    "Lot":        lot,
                    "LotType":    lot_type,
                    "Sample":     sample,
                    "Rep":        int(rep_row["Rep"]),
                    "Result":     rep_row["Result"],
                    "IsRerun":    is_rerun,
                })

            # Also store original rows as superseded (for display)
            for sup_ver in superseded_vers:
                sup_df = sample_versions[sup_ver]
                sup_valid = sup_df[(sup_df["Flag"] == 1) &
                                   (~sup_df["BelowCurve"])]
                sup_mean = sup_valid["Result"].mean() if len(sup_valid) else None
                sup_rlu  = sup_df[sup_df["Flag"]==1]["RLU"].mean() if len(sup_df) else None
                summary_rows.append({
                    "Study":        study_name,
                    "Timepoint":    timepoint,
                    "Instrument":   instrument,
                    "RunNum":       run_num,
                    "RunID":        f"Run{run_num} (Original — superseded)",
                    "Lot":          lot,
                    "LotType":      lot_type,
                    "Sample":       sample,
                    "RunDate":      run_date,
                    "Reps":         f"{len(sup_valid)} / {len(sup_df)}",
                    "RLU_Avg":      round(sup_rlu,  0) if sup_rlu  is not None else None,
                    "Dose_Avg":     round_conc(sup_mean) if sup_mean is not None else None,
                    "RLU_CV_pct":   None,
                    "Dose_CV_pct":  None,
                    "ValidReps":    len(sup_valid),
                    "TotalReps":    len(sup_df),
                    "MeanDose":     sup_mean,
                    "IsRerun":      True,
                    "AcceptedVer":  -1,  # -1 = superseded, excluded from calcs
                })

    summary_df = pd.DataFrame(summary_rows)
    reps_df    = pd.DataFrame(rep_rows)
    return summary_df, reps_df


# ─────────────────────────────────────────────
# LOAD ENTIRE STUDY
# ─────────────────────────────────────────────

def load_study(study_folder):
    """
    Load all timepoints from a study folder.
    Returns (summary_df, reps_df) across all timepoints.
    """
    study_name = os.path.basename(study_folder.rstrip("/\\"))
    # Convert HE4_Stab11 → HE4 Stab 11
    # HE4_Stab11 → HE4 Stab 11
    # Replace underscores, then insert space between letters and digits in "Stab##"
    study_display = re.sub(r'(?i)(Stab)(\d+)', r'\1 \2', study_name.replace("_", " "))

    # Load config — sets specs, CV limit, sample list
    assay_name, study_number, config_loaded = load_config(study_folder)
    # If config loaded, use its assay+study name for display
    if config_loaded:
        study_display = f"{assay_name} {study_number}"

    # Find TP subfolders
    tp_folders = {}
    for item in os.listdir(study_folder):
        full = os.path.join(study_folder, item)
        if os.path.isdir(full) and re.match(r'^TP\d+$', item, re.IGNORECASE):
            tp_num = int(re.sub(r'[^0-9]', '', item))
            tp_folders[tp_num] = (item.upper(), full)

    if not tp_folders:
        print(f"❌ No TP subfolders found in {study_folder}")
        return pd.DataFrame(), pd.DataFrame()

    all_summary, all_reps = [], []

    for tp_num in sorted(tp_folders.keys()):
        tp_label, tp_path = tp_folders[tp_num]
        print(f"  Loading {tp_label}...")
        s_df, r_df = load_timepoint(tp_path, tp_label, study_display)
        if not s_df.empty:
            all_summary.append(s_df)
        if not r_df.empty:
            all_reps.append(r_df)
        print(f"    → {len(s_df)} run-sample records, "
              f"{len(r_df)} replicates")

    summary_df = pd.concat(all_summary, ignore_index=True) if all_summary else pd.DataFrame()
    reps_df    = pd.concat(all_reps,    ignore_index=True) if all_reps    else pd.DataFrame()
    return study_display, summary_df, reps_df



# ─────────────────────────────────────────────
# PASS / FAIL LOGIC
# ─────────────────────────────────────────────

def per_run_pass(row):
    """Evaluate a single run-sample record against specs."""
    sample    = row["Sample"]
    mean      = row["MeanDose"]
    n_valid   = row["ValidReps"]
    n_total   = row["TotalReps"]
    dose_cv   = row["Dose_CV_pct"]
    rlu_cv    = row.get("RLU_CV_pct")
    is_val    = row["LotType"] == "Validity"
    lot       = row.get("Lot", "Lot1")
    timepoint = row.get("Timepoint", "TP1")

    if row["AcceptedVer"] == -1:
        return "SUPERSEDED"
    if mean is None:
        return "FAIL"

    indiv_spec, mean_spec, validity_spec = get_specs_for(sample, lot, timepoint)

    if is_val and sample.startswith("RefCon"):
        spec = validity_spec or {}
        return "PASS" if spec and spec["LSL"] <= mean <= spec["USL"] else "FAIL"

    spec = mean_spec or {}
    if not spec:
        return "N/A"

    within_cv, _ = get_cv_limits(sample, lot)
    # Within-run validity uses RLU %CV per protocol
    cv_ok   = rlu_cv is not None and rlu_cv <= within_cv
    rep_ok  = n_valid >= MIN_REPS and n_valid == n_total
    spec_ok = spec["LSL"] <= mean <= spec["USL"]
    return "PASS" if (cv_ok and rep_ok and spec_ok) else "FAIL"


def rep_color(value, sample, lot="Lot1", timepoint="TP1"):
    if value is None or pd.isna(value):
        return None
    indiv_spec, _, _ = get_specs_for(sample, lot, timepoint)
    if not indiv_spec:
        return None
    return "PASS" if indiv_spec["LSL"] <= value <= indiv_spec["USL"] else "FAIL"


# ─────────────────────────────────────────────
# EXCEL STYLE HELPERS
# ─────────────────────────────────────────────

def conc_fmt():
    """Return Excel number format string for concentration values."""
    return "0." + "0" * CONC_DECIMALS if CONC_DECIMALS > 0 else "0"

def round_conc(value):
    """Round a concentration value to the configured decimal places."""
    if value is None:
        return None
    try:
        return round(float(value), CONC_DECIMALS)
    except (TypeError, ValueError):
        return value
# ─────────────────────────────────────────────
# EXCEL STYLE HELPERS
# ─────────────────────────────────────────────

C_HEADER_BG  = "1F3864"
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "2E75B6"
C_SECTION_FG = "FFFFFF"
C_SUB_BG     = "4472C4"
C_PASS       = "C6EFCE"
C_FAIL       = "FFC7CE"
C_SUPER      = "F2F2F2"   # light grey for superseded rows
C_PASS_FONT  = "006100"
C_FAIL_FONT  = "9C0006"
C_ALT_ROW    = "F2F7FC"

thin = Side(style="thin",   color="B8CCE4")
med  = Side(style="medium", color="1F3864")

def bdr(thick=False):
    s = med if thick else thin
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row, cols, labels, bg=C_HEADER_BG, fg=C_HEADER_FG):
    for col, label in zip(cols, labels):
        c = ws.cell(row=row, column=col, value=label)
        c.font      = Font(bold=True, color=fg, name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = bdr()

def wdc(ws, row, col, value, pf=None, alt=False, fmt=None,
        bold=False, superseded=False):
    c = ws.cell(row=row, column=col, value=value)
    if superseded:
        fg_color = "888888"
        fill     = PatternFill("solid", fgColor=C_SUPER)
    elif pf == "PASS":
        fg_color = C_PASS_FONT
        fill     = PatternFill("solid", fgColor=C_PASS)
    elif pf == "FAIL":
        fg_color = C_FAIL_FONT
        fill     = PatternFill("solid", fgColor=C_FAIL)
    elif alt:
        fg_color = "000000"
        fill     = PatternFill("solid", fgColor=C_ALT_ROW)
    else:
        fg_color = "000000"
        fill     = PatternFill("solid", fgColor="FFFFFF")
    c.font      = Font(name="Arial", size=10, bold=bold, color=fg_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = bdr()
    c.fill      = fill
    if fmt and not isinstance(value, str):
        c.number_format = fmt

def section_heading(ws, row, c1, c2, title, bg=C_SECTION_BG):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row, end_column=c2)
    # Style every cell in the merge range — prevents visible border lines
    # from appearing between unformatted cells within the merged area
    fill   = PatternFill("solid", fgColor=bg)
    no_bdr = Border(
        left   = Side(style="thin", color=bg),
        right  = Side(style="thin", color=bg),
        top    = Side(style="thin", color=bg),
        bottom = Side(style="thin", color=bg),
    )
    for col in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.border    = no_bdr
    # Title cell gets the actual content and visible outer border
    c = ws.cell(row=row, column=c1, value=title)
    c.font      = Font(bold=True, color=C_SECTION_FG, name="Arial", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = bdr(thick=True)
    ws.row_dimensions[row].height = 20

def set_row_heights(ws):
    """Apply consistent row heights across a sheet."""
    for row_cells in ws.iter_rows():
        r = row_cells[0].row
        has_wrap   = any(c.alignment and c.alignment.wrap_text and
                         c.value and "\n" in str(c.value)
                         for c in row_cells)
        is_title   = any(c.font and c.font.size and c.font.size >= 14
                         for c in row_cells if c.value)
        is_section = any(c.font and c.font.size and c.font.size >= 11
                         for c in row_cells if c.value)
        if is_title:
            ws.row_dimensions[r].height = 28
        elif has_wrap:
            ws.row_dimensions[r].height = 42
        elif is_section:
            ws.row_dimensions[r].height = 20
        elif any(c.value for c in row_cells):
            ws.row_dimensions[r].height = 16


# ─────────────────────────────────────────────
# REPLICATE GRID
# ─────────────────────────────────────────────

def write_rep_grid(ws, start_row, start_col, sample, reps_df, lot_type,
                   single_rep=False):
    """
    Write individual replicate grid for one sample/lot.
    Columns: one per run (grouped by instrument).
    Rows: Rep 1-12, then Std Dev, Mean, %CV, Valid?, Acceptance.
    Returns next available row number.
    """
    instruments = sorted(reps_df["Instrument"].unique())
    run_labels  = []
    for inst in instruments:
        inst_runs = sorted(
            reps_df[reps_df["Instrument"] == inst]["RunID"].unique()
        )
        for rid in inst_runs:
            run_labels.append((inst, rid))

    n_runs     = len(run_labels)
    # Get specs from the first available lot/timepoint in the reps dataframe
    _lot_rg = reps_df["Lot"].iloc[0] if not reps_df.empty else "Lot1"
    _tp_rg  = reps_df["Timepoint"].iloc[0] if "Timepoint" in reps_df.columns and not reps_df.empty else "TP1"
    spec_indiv_t, spec_mean_t, spec_val_t = get_specs_for(sample, _lot_rg, _tp_rg)
    spec_indiv = spec_indiv_t or INDIV_REP_SPECS.get(sample, {})
    spec_mean  = (spec_val_t if lot_type == "Validity"
                  else spec_mean_t) or (
                  VALIDITY_SPECS if lot_type == "Validity" else MEAN_SPECS).get(sample, {})

    row = start_row

    # ── Instrument header (merged per instrument) ──
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row + 1, end_column=start_col)
    c = ws.cell(row=row, column=start_col, value="Rep")
    c.font      = Font(bold=True, name="Arial", size=10, color=C_HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = bdr()

    inst_groups = []
    for inst in instruments:
        idxs = [i for i, (ins, _) in enumerate(run_labels) if ins == inst]
        inst_groups.append((inst, idxs[0], idxs[-1]))

    for inst, first, last in inst_groups:
        c1 = start_col + 1 + first
        c2 = start_col + 1 + last
        if c1 < c2:
            ws.merge_cells(start_row=row, start_column=c1,
                           end_row=row, end_column=c2)
        # Fill all cells in merge to prevent visible split lines
        hdr_fill = PatternFill("solid", fgColor=C_HEADER_BG)
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = hdr_fill
            cell.border = bdr()
        c = ws.cell(row=row, column=c1, value=inst)
        c.font      = Font(bold=True, name="Arial", size=10, color=C_HEADER_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(c1 + 1, c2 + 1):
            cc = ws.cell(row=row, column=col)
            cc.fill   = PatternFill("solid", fgColor=C_HEADER_BG)
            cc.border = bdr()
    row += 1

    # ── Run ID sub-header ──
    for i, (inst, rid) in enumerate(run_labels):
        short = rid.replace("HE4", "").replace("ST11", "").replace("TP1", "")
        c = ws.cell(row=row, column=start_col + 1 + i, value=rid)
        c.font      = Font(bold=True, name="Arial", size=9, color=C_HEADER_FG)
        c.fill      = PatternFill("solid", fgColor=C_SUB_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = bdr()
        ws.row_dimensions[row].height = 30
    row += 1

    # ── Replicate rows ──
    if single_rep:
        reps_df = reps_df[reps_df["Rep"] == 1].copy()
    max_reps = int(reps_df["Rep"].max()) if not reps_df.empty else 1
    max_reps = max(max_reps, 12) if not single_rep else max_reps

    for rep_num in range(1, max_reps + 1):
        alt = (rep_num % 2 == 0)
        c = ws.cell(row=row, column=start_col, value=rep_num)
        c.font      = Font(bold=True, name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid",
                                  fgColor=C_ALT_ROW if alt else "FFFFFF")
        c.border    = bdr()

        for i, (inst, rid) in enumerate(run_labels):
            match = reps_df[(reps_df["Instrument"] == inst) &
                            (reps_df["RunID"]      == rid)  &
                            (reps_df["Rep"]        == rep_num)]
            has_data = not match.empty and pd.notna(match["Result"].iloc[0])
            val = round(match["Result"].iloc[0], 1) if has_data else "N/A"
            lot_for_rep = reps_df["Lot"].iloc[0] if not reps_df.empty else "Lot1"
            tp_for_rep  = reps_df["Timepoint"].iloc[0] if "Timepoint" in reps_df.columns and not reps_df.empty else "TP1"
            pf  = rep_color(val, sample, lot_for_rep, tp_for_rep) if has_data else None
            wdc(ws, row, start_col + 1 + i, val, pf=pf, fmt="0.0")
        row += 1

    # ── Stats rows ──
    stat_labels = (["Mean", "Accepted"] if single_rep
                    else ["Std Dev", "Mean", "%CV", "Valid?", "Accepted"])
    for label in stat_labels:
        c = ws.cell(row=row, column=start_col, value=label)
        c.font      = Font(bold=True, name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor="D9E1F2")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()

        for i, (inst, rid) in enumerate(run_labels):
            col_vals = reps_df[(reps_df["Instrument"] == inst) &
                               (reps_df["RunID"]      == rid)
                               ]["Result"].dropna()
            mn = col_vals.mean() if len(col_vals) else None
            sd = col_vals.std()  if len(col_vals) >= 2 else None
            cv = (sd / mn * 100) if (mn and sd is not None) else None

            if label == "Std Dev":
                wdc(ws, row, start_col + 1 + i,
                    round(sd, 2) if sd is not None else None, fmt="0.00")
            elif label == "Mean":
                wdc(ws, row, start_col + 1 + i,
                    round_conc(mn) if mn is not None else None, fmt=conc_fmt())
            elif label == "%CV":
                # Dose %CV — informational only, no pass/fail spec
                wdc(ws, row, start_col + 1 + i,
                    round(cv, 1) / 100 if cv is not None else None,
                    pf=None, fmt="0.0%")
            elif label == "Valid?":
                # Valid = all individual reps in this run are within indiv spec
                run_reps = col_vals.dropna().tolist()
                if spec_indiv and len(run_reps) > 0:
                    all_valid = all(spec_indiv["LSL"] <= r <= spec_indiv["USL"] for r in run_reps if r is not None)
                    v = "Valid" if all_valid else "No"
                else:
                    v = "Valid"  # no spec defined, assume valid
                wdc(ws, row, start_col + 1 + i, v,
                    pf="PASS" if v == "Valid" else "FAIL")
            elif label in ("Acceptance", "Accepted"):
                if mn is not None and spec_mean:
                    v  = "YES" if spec_mean["LSL"] <= mn <= spec_mean["USL"] else "No"
                    wdc(ws, row, start_col + 1 + i, v,
                        pf="PASS" if v == "YES" else "FAIL")
                else:
                    wdc(ws, row, start_col + 1 + i, None)

    # ── Overall stats row — compact 2-column table ──────────────────
    if not single_rep:
        row += 1  # blank spacer row

        import numpy as np

        all_vals = reps_df["Result"].dropna().tolist()

        if all_vals:
            mn  = float(np.mean(all_vals))
            sd  = float(np.std(all_vals, ddof=1)) if len(all_vals) > 1 else 0.0
            cv  = (sd / mn * 100) if mn else 0.0
            vld = all(spec_indiv["LSL"] <= v <= spec_indiv["USL"]
                      for v in all_vals) if spec_indiv else None
            acc = (spec_mean["LSL"] <= mn <= spec_mean["USL"]
                   if spec_mean else None)

            # Two columns only: start_col (label) and start_col+1 (value)
            OV_END = start_col + 1
            no_border = Border()

            # Clear all run columns beyond col 2 for entire overall block height
            clear_rows = 7  # Overall heading + 5 stat rows
            for r_off in range(clear_rows):
                for gap_col in range(start_col + 2, start_col + n_runs + 1):
                    gc = ws.cell(row=row + r_off, column=gap_col)
                    gc.value  = None
                    gc.fill   = PatternFill("solid", fgColor="FFFFFF")
                    gc.border = no_border

            # "Overall" heading — merged across both columns, centred
            ws.merge_cells(start_row=row, start_column=start_col,
                           end_row=row,   end_column=OV_END)
            ov_fill = PatternFill("solid", fgColor="365F91")
            for col in range(start_col, OV_END + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill   = ov_fill
                cell.border = bdr()
            c = ws.cell(row=row, column=start_col, value="Overall")
            c.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 16
            row += 1

            for label, val, pf, fmt_key in [
                ("Std Dev", round(sd, 2),       None,  "0.00"),
                ("Mean",    round_conc(mn),      None,  conc_fmt()),
                ("%CV",     round(cv, 1) / 100,  None,  "0.0%"),
                ("Valid?",
                 "Valid" if vld else ("Invalid" if vld is not None else "N/A"),
                 "PASS" if vld else ("FAIL" if vld is not None else None), None),
                ("Accepted",
                 "YES" if acc else ("No" if acc is not None else "N/A"),
                 "PASS" if acc else ("FAIL" if acc is not None else None), None),
            ]:
                # Label cell (col 1)
                c = ws.cell(row=row, column=start_col, value=label)
                c.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                c.fill      = PatternFill("solid", fgColor="2E4E7E")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

                # Value cell (col 2) — no merge, single cell
                wdc(ws, row, OV_END, val, pf=pf, fmt=fmt_key)

                ws.row_dimensions[row].height = 16
                row += 1

    return row







# ─────────────────────────────────────────────
# PER-TIMEPOINT REPORT BUILDER
# ─────────────────────────────────────────────

def build_tp_report(summary_df, reps_df, study_display, timepoint, out_path):
    """Build the per-timepoint Excel report."""

    tp_summary = summary_df[summary_df["Timepoint"] == timepoint].copy()
    tp_reps    = reps_df[reps_df["Timepoint"]    == timepoint].copy()

    tp_summary["PassFail"] = tp_summary.apply(per_run_pass, axis=1)

    wb = Workbook()

    # ══ SUMMARY SHEET ══════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.row_dimensions[1].height = 8

    # Title banner
    ws.merge_cells("B2:L3")
    t = ws["B2"]
    t.value     = f"{study_display} — {timepoint} Summary Report"
    t.font      = Font(bold=True, size=16, color="FFFFFF", name="Arial")
    t.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28

    # Metadata
    instruments = sorted(tp_summary["Instrument"].unique())
    lots        = sorted(tp_summary["Lot"].unique())
    run_date    = tp_summary["RunDate"].dropna().iloc[0] if not tp_summary["RunDate"].dropna().empty else "N/A"

    meta = [
        ("Study",       study_display),
        ("Timepoint",   timepoint),
        ("Instruments", ", ".join(instruments)),
        ("Lots",        ", ".join(lots)),
        ("Run Date",    run_date),
    ]
    for i, (k, v) in enumerate(meta):
        r = 5 + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=10)
        ws.cell(row=r, column=2, value=k + ":").font = Font(
            bold=True, name="Arial", size=10)
        ws.cell(row=r, column=4, value=v).font = Font(
            name="Arial", size=10)

    row = 12

    feas_df = tp_summary[tp_summary["LotType"] == "Feasibility"]
    val_df  = tp_summary[tp_summary["LotType"] == "Validity"]

    # ── Feasibility lots ──
    for lot in sorted(feas_df["Lot"].unique()):
        lot_df = feas_df[feas_df["Lot"] == lot].copy()

        section_heading(ws, row, 2, 12, f"{lot} — Run Summary")
        row += 1
        apply_header(ws, row, range(2, 13),
                     ["Instrument", "Run ID", "Sample", "Reps",
                      "RLU Avg", "Dose Avg\n(pmol/L)", "RLU %CV",
                      "Dose %CV", "Mean\nSpec Low", "Mean\nSpec High",
                      "Result"])
        row += 1

        alt = False
        for _, r in lot_df.sort_values(
                ["Instrument", "RunNum", "Sample"]).iterrows():
            _, mean_spec_r, _ = get_specs_for(r["Sample"], r["Lot"], r["Timepoint"])
            spec       = mean_spec_r or MEAN_SPECS.get(r["Sample"], {})
            pf         = r["PassFail"]
            superseded = (r["AcceptedVer"] == -1)
            note       = "⚠ Superseded by rerun" if superseded else (
                         "↺ Rerun" if r["IsRerun"] and not superseded else "")
            vals = [r["Instrument"], r["RunID"], r["Sample"], r["Reps"],
                    r["RLU_Avg"], r["Dose_Avg"],
                    r["RLU_CV_pct"] / 100 if r["RLU_CV_pct"] is not None else "N/A",
                    r["Dose_CV_pct"] / 100 if r["Dose_CV_pct"] is not None else "N/A",
                    spec.get("LSL"), spec.get("USL"),
                    pf if not superseded else "—"]
            fmts = [None, None, None, None, "#,##0", "0.0",
                    "0.0%", "0.0%", "0.0", "0.0", None]
            for ci, (v, fmt) in enumerate(zip(vals, fmts)):
                pf_arg = (pf if ci == 10 and not superseded else None)
                wdc(ws, row, ci + 2, v, pf=pf_arg, alt=alt,
                    fmt=fmt, superseded=superseded)
            alt = not alt
            row += 1

        # Between-run RefCon validity
        row += 1
        section_heading(ws, row, 2, 12,
                        f"{lot} — Between-Run Validity (RefCon Grand Mean)",
                        bg="365F91")
        row += 1
        apply_header(ws, row, range(2, 13),
                     ["Sample", "Instrument", "Run1\nMean", "Run2\nMean",
                      "Run3\nMean", "Grand\nMean", "SD", "%CV",
                      "Mean\nSpec Low", "Mean\nSpec High", "Result"],
                     bg=C_SUB_BG, fg="000000")
        row += 1

        active_lot = lot_df[lot_df["AcceptedVer"] >= 0]
        for samp in ["RefCon1", "RefCon2", "RefCon3"]:
            for inst in sorted(active_lot["Instrument"].unique()):
                inst_lot = lot  # lot variable from outer loop
                tp = timepoint   # timepoint variable from function param
                sub = active_lot[
                    (active_lot["Sample"]     == samp) &
                    (active_lot["Instrument"] == inst)
                ].sort_values("RunNum")
                means   = sub["MeanDose"].dropna().tolist()
                rm      = means[:3] + [None] * (3 - len(means[:3]))
                valid_m = [m for m in rm if m is not None]
                grand   = round_conc(sum(valid_m) / len(valid_m)) if valid_m else None
                sd      = round(pd.Series(valid_m).std(), 2) \
                          if len(valid_m) >= 2 else None
                cv      = round(sd / grand, 4) \
                          if (sd is not None and grand) else None
                _, mean_spec_br, _ = get_specs_for(samp, inst_lot, tp)
                spec    = mean_spec_br or MEAN_SPECS.get(samp, {})
                acc     = ("PASS" if spec and grand and
                           spec["LSL"] <= grand <= spec["USL"]
                           else "FAIL" if grand else None)
                rm_r = [round(m, 1) if m is not None else None for m in rm]
                rv = [samp, inst, rm_r[0], rm_r[1], rm_r[2], grand,
                      sd if sd is not None else 0.0,
                      cv if cv is not None else 0.0,
                      spec.get("LSL"), spec.get("USL"), acc]
                fmts = [None, None, conc_fmt(), conc_fmt(), conc_fmt(), conc_fmt(),
                        "0.00", "0.0%", conc_fmt(), conc_fmt(), None]
                for ci, (v, fmt) in enumerate(zip(rv, fmts)):
                    wdc(ws, row, ci + 2, v,
                        pf=(_get_br_cv_pf(rv[7], samp) if ci == 7 else acc if ci == 10 else None), fmt=fmt)
                row += 1
        row += 2

    # ── Validity lot ──
    if not val_df.empty:
        active_val = val_df[val_df["AcceptedVer"] >= 0].copy()

        section_heading(ws, row, 2, 9,
                        "Validity — Single-Replicate RefCon Runs")
        row += 1
        apply_header(ws, row, range(2, 10),
                     ["Instrument", "Run ID", "Sample",
                      "RLU", "Result\n(pmol/L)",
                      "Spec Low", "Spec High", "Result"])
        row += 1

        alt = False
        for _, r in active_val.sort_values(
                ["Instrument", "RunNum", "Sample"]).iterrows():
            _, _, val_spec = get_specs_for(r["Sample"], r["Lot"], r["Timepoint"])
            spec       = val_spec or VALIDITY_SPECS.get(r["Sample"], {})
            pf         = r["PassFail"]
            superseded = (r["AcceptedVer"] == -1)
            note       = "↺ Rerun" if r["IsRerun"] and not superseded else ""
            vals = [r["Instrument"], r["RunID"], r["Sample"],
                    r["RLU_Avg"], r["Dose_Avg"],
                    spec.get("LSL"), spec.get("USL"),
                    pf if not superseded else "—"]
            fmts = [None, None, None, "#,##0", conc_fmt(),
                    conc_fmt(), conc_fmt(), None]
            for ci, (v, fmt) in enumerate(zip(vals, fmts)):
                wdc(ws, row, ci + 2, v,
                    pf=(pf if ci == 7 and not superseded else None),
                    alt=alt, fmt=fmt, superseded=superseded)
            alt = not alt
            row += 1

        # Grand mean acceptance
        row += 1
        section_heading(ws, row, 2, 9,
                        "Validity — Grand Mean Acceptance", bg="365F91")
        row += 1
        apply_header(ws, row, range(2, 10),
                     ["Sample", "# Values", "Grand Mean\n(pmol/L)",
                      "Spec Low", "Spec High", "Min", "Max", "Result"],
                     bg=C_SUB_BG, fg="000000")
        row += 1

        for samp in ["RefCon1", "RefCon2", "RefCon3"]:
            sub   = active_val[active_val["Sample"] == samp]["MeanDose"].dropna()
            grand = round(sub.mean(), 1) if len(sub) else None
            _, _, val_spec_gm = get_specs_for(samp, "Lot1", tp)
            spec  = val_spec_gm or VALIDITY_SPECS.get(samp, {})
            acc   = ("PASS" if spec and grand and
                     spec["LSL"] <= grand <= spec["USL"]
                     else "FAIL" if grand else None)
            rv = [samp, len(sub), round_conc(grand),
                  spec.get("LSL"), spec.get("USL"),
                  round_conc(sub.min()) if len(sub) else None,
                  round_conc(sub.max()) if len(sub) else None, acc]
            fmts = [None, None, conc_fmt(), conc_fmt(), conc_fmt(),
                    conc_fmt(), conc_fmt(), None]
            for ci, (v, fmt) in enumerate(zip(rv, fmts)):
                wdc(ws, row, ci + 2, v,
                    pf=(acc if ci == 7 else None), fmt=fmt)
            row += 1
        row += 2

    # Summary column widths and row heights
    for col, w in {"A": 2, "B": 13, "C": 10, "D": 12, "E": 10,
                   "F": 10, "G": 12, "H": 9, "I": 9,
                   "J": 12, "K": 12, "L": 12}.items():
        ws.column_dimensions[col].width = w
    set_row_heights(ws)

    # ══ PER-SAMPLE DETAIL SHEETS ═══════════════════════════════════════
    for samp in SAMPLE_ORDER:
        sdf  = tp_summary[tp_summary["Sample"] == samp].copy()
        srep = tp_reps[tp_reps["Sample"] == samp].copy()
        if sdf.empty:
            continue

        ws2 = wb.create_sheet(samp)
        ws2.sheet_view.showGridLines = False
        ws2.column_dimensions["A"].width = 2

        rw = 1
        section_heading(ws2, rw, 2, 14,
                        f"{samp} — {timepoint} Run Summary (All Lots)")
        rw += 1
        apply_header(ws2, rw, range(2, 15),
                     ["Lot", "Lot Type", "Instrument", "Run ID", "Reps",
                      "RLU Avg", "Dose Avg\n(pmol/L)", "Dose %CV",
                      "Indiv Rep\nSpec Low", "Indiv Rep\nSpec High",
                      "Mean\nSpec Low", "Mean\nSpec High",
                      "Pass/Fail"])
        rw += 1

        # Use TP1 specs as display reference for the summary table
        spec_i, spec_m, _ = get_specs_for(samp, "Lot1", tp)
        spec_i = spec_i or INDIV_REP_SPECS.get(samp, {})
        spec_m = spec_m or MEAN_SPECS.get(samp, {})
        alt    = False
        for _, r in sdf.sort_values(
                ["Lot", "Instrument", "RunNum"]).iterrows():
            pf         = r["PassFail"]
            superseded = (r["AcceptedVer"] == -1)
            note = "⚠ Superseded by rerun" if superseded else (
                   "↺ Rerun" if r["IsRerun"] and not superseded else "")
            vals = [r["Lot"], r["LotType"], r["Instrument"], r["RunID"],
                    r["Reps"], r["RLU_Avg"], r["Dose_Avg"],
                    r["Dose_CV_pct"] / 100 if r["Dose_CV_pct"] is not None else "N/A",
                    spec_i.get("LSL"), spec_i.get("USL"),
                    spec_m.get("LSL"), spec_m.get("USL"),
                    pf if not superseded else "—"]
            fmts = [None, None, None, None, None, "#,##0", conc_fmt(),
                    "0.0%", conc_fmt(), conc_fmt(), conc_fmt(), conc_fmt(), None]
            for ci, (v, fmt) in enumerate(zip(vals, fmts)):
                wdc(ws2, rw, ci + 2, v,
                    pf=(pf if ci == 12 and not superseded else None),
                    alt=alt, fmt=fmt, superseded=superseded)
            alt = not alt
            rw += 1

        rw += 1

        # Replicate grids per feasibility lot
        for lot in sorted(srep[srep["LotType"] == "Feasibility"]["Lot"].unique()):
            lot_reps = srep[srep["Lot"] == lot]
            if lot_reps.empty:
                continue
            n_runs = len(lot_reps[["Instrument","RunID"]].drop_duplicates())
            section_heading(ws2, rw, 2, 2 + n_runs,
                            f"{lot} — Individual Replicate Values  "
                            f"(Indiv Spec: "
                            f"{spec_i.get('LSL', '')}–"
                            f"{spec_i.get('USL', '')} pmol/L)",
                            bg="365F91")
            rw += 1
            rw = write_rep_grid(ws2, rw, 2, samp, lot_reps, "Feasibility")
            rw += 1


        # Validity replicate grid — RefCon samples only
        # Validity lot runs one rep per RefCon per run by design.
        # Panel samples are not included in the validity lot.
        val_reps = srep[srep["LotType"] == "Validity"]
        if not val_reps.empty and samp.startswith("RefCon"):
            n_runs = len(val_reps[["Instrument","RunID"]].drop_duplicates())
            _, _, spec_v = get_specs_for(samp, "Lot1", "TP1")
            spec_v = spec_v or VALIDITY_SPECS.get(samp, {})
            spec_range = (
                f"{spec_v.get('LSL','')}–{spec_v.get('USL','')} pmol/L"
                if spec_v else ""
            )
            section_heading(ws2, rw, 2, 2 + n_runs,
                            f"Validity — Single-Replicate Values"
                            + (f"  (Spec: {spec_range})" if spec_range else ""),
                            bg="365F91")
            rw += 1
            rw = write_rep_grid(ws2, rw, 2, samp, val_reps, "Validity",
                                single_rep=True)
            rw += 1
        # Column widths
        detail_widths = {
            "A": 2,  "B": 10, "C": 12, "D": 12, "E": 10,
            "F": 9,  "G": 10, "H": 12, "I": 10, "J": 12,
            "K": 12, "L": 12, "M": 12,
        }
        for ci in range(15, 22):
            detail_widths[get_column_letter(ci)] = 10
        for col, w in detail_widths.items():
            ws2.column_dimensions[col].width = w
        set_row_heights(ws2)

    wb.save(out_path)
    print(f"  ✅ {timepoint} report saved: {out_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────



# ─────────────────────────────────────────────
# TABLEAU CSV EXPORT
# ─────────────────────────────────────────────

def export_tableau_csv(study_folder, study_display, summary_df, reps_df):
    """
    Export two clean CSV files for Tableau:

    1. {Study}_Tableau_RunSummary.csv
       One row per (Timepoint, Lot, Instrument, Run, Sample)
       Contains: grand means, CVs, pass/fail, spec limits
       Use for: trend charts, run-level analysis

    2. {Study}_Tableau_Replicates.csv
       One row per individual replicate
       Contains: raw result values, pass/fail vs indiv spec
       Use for: replicate-level drill-down, box plots
    """
    active = summary_df[summary_df["AcceptedVer"] >= 0].copy()
    active["PassFail"] = active.apply(per_run_pass, axis=1)

    # ── Run summary CSV ──────────────────────────────────────────────
    run_rows = []
    for _, r in active.iterrows():
        samp      = r["Sample"]
        lot_type  = r["LotType"]
        spec_mean = (VALIDITY_SPECS if lot_type == "Validity"
                     else MEAN_SPECS).get(samp, {})
        spec_indiv = INDIV_REP_SPECS.get(samp, {})

        run_rows.append({
            "Study":            study_display,
            "Timepoint":        r["Timepoint"],
            "Timepoint_Num":    int(re.sub(r"[^0-9]", "", r["Timepoint"])),
            "Lot":              r["Lot"],
            "Lot_Type":         r["LotType"],
            "Instrument":       r["Instrument"],
            "Run_ID":           r["RunID"],
            "Run_Num":          r["RunNum"],
            "Sample":           samp,
            "Sample_Type":      ("Panel" if samp.startswith("Panel") else
                                 "RefCon" if samp.startswith("RefCon") else "Other"),
            "Run_Date":         r["RunDate"],
            "Valid_Reps":       r["ValidReps"],
            "Total_Reps":       r["TotalReps"],
            "RLU_Avg":          r["RLU_Avg"],
            "Mean_Conc":        r["Dose_Avg"],
            "RLU_CV_pct":       round(r["RLU_CV_pct"], 4) if r["RLU_CV_pct"] is not None else None,
            "Dose_CV_pct":      round(r["Dose_CV_pct"], 4) if r["Dose_CV_pct"] is not None else None,
            "Mean_Spec_LSL":    spec_mean.get("LSL"),
            "Mean_Spec_USL":    spec_mean.get("USL"),
            "Indiv_Spec_LSL":   spec_indiv.get("LSL"),
            "Indiv_Spec_USL":   spec_indiv.get("USL"),
            "Pass_Fail":        r["PassFail"],
            "Is_Rerun":         r["IsRerun"],
        })

    run_df = pd.DataFrame(run_rows)

    # Grand mean per (Timepoint, Lot, Sample) — useful for trend charts
    gm = (run_df.groupby(["Study","Timepoint","Timepoint_Num","Lot",
                           "Lot_Type","Sample","Sample_Type",
                           "Mean_Spec_LSL","Mean_Spec_USL",
                           "Indiv_Spec_LSL","Indiv_Spec_USL"])
               .agg(Grand_Mean=("Mean_Conc","mean"),
                    Grand_Mean_CV=("Mean_Conc", lambda x:
                                   round(x.std()/x.mean()*100, 4)
                                   if len(x)>=2 and x.mean() else None),
                    Runs_Passing=("Pass_Fail", lambda x: (x=="PASS").sum()),
                    Total_Runs=("Pass_Fail", "count"))
               .reset_index())
    gm["Grand_Mean"] = gm["Grand_Mean"].round(2)
    gm["Grand_Mean_Pass"] = gm.apply(
        lambda r: ("PASS" if r["Mean_Spec_LSL"] and r["Mean_Spec_USL"] and
                   r["Mean_Spec_LSL"] <= r["Grand_Mean"] <= r["Mean_Spec_USL"]
                   else "FAIL"), axis=1)

    # Save all Tableau files into a Tableau/ subfolder
    tableau_folder = os.path.join(study_folder, "Tableau")
    os.makedirs(tableau_folder, exist_ok=True)

    run_path = os.path.join(
        tableau_folder,
        f"{study_display.replace(' ', '_')}_Tableau_RunSummary.csv"
    )
    grand_path = os.path.join(
        tableau_folder,
        f"{study_display.replace(' ', '_')}_Tableau_GrandMeans.csv"
    )
    run_df.to_csv(run_path,   index=False)
    gm.to_csv(grand_path,     index=False)

    # ── Replicates CSV ───────────────────────────────────────────────
    rep_rows = []
    for _, r in reps_df[reps_df["AcceptedVer"] >= 0].iterrows() if "AcceptedVer" in reps_df.columns else reps_df.iterrows():
        samp       = r["Sample"]
        spec_indiv = INDIV_REP_SPECS.get(samp, {})
        result     = r["Result"]
        rep_pf     = ("PASS" if spec_indiv and result is not None and
                      not pd.isna(result) and
                      spec_indiv["LSL"] <= result <= spec_indiv["USL"]
                      else "FAIL" if result is not None and not pd.isna(result)
                      else None)
        rep_rows.append({
            "Study":          study_display,
            "Timepoint":      r["Timepoint"],
            "Timepoint_Num":  int(re.sub(r"[^0-9]", "", r["Timepoint"])),
            "Lot":            r["Lot"],
            "Lot_Type":       r["LotType"],
            "Instrument":     r["Instrument"],
            "Run_ID":         r["RunID"],
            "Run_Num":        r["RunNum"],
            "Sample":         samp,
            "Sample_Type":    ("Panel" if samp.startswith("Panel") else
                               "RefCon" if samp.startswith("RefCon") else "Other"),
            "Rep_Num":        r["Rep"],
            "Result":         round(result, 2) if result is not None and not pd.isna(result) else None,
            "Indiv_Spec_LSL": spec_indiv.get("LSL"),
            "Indiv_Spec_USL": spec_indiv.get("USL"),
            "Rep_Pass_Fail":  rep_pf,
            "Is_Rerun":       r["IsRerun"],
        })

    rep_df  = pd.DataFrame(rep_rows)
    rep_path = os.path.join(
        tableau_folder,
        f"{study_display.replace(' ', '_')}_Tableau_Replicates.csv"
    )
    rep_df.to_csv(rep_path, index=False)

    print(f"  ✅ Tableau CSVs saved to Tableau/ subfolder:")
    print(f"     Run summary:  {os.path.basename(run_path)}  ({len(run_df)} rows)")
    print(f"     Grand means:  {os.path.basename(grand_path)}  ({len(gm)} rows)")
    print(f"     Replicates:   {os.path.basename(rep_path)}  ({len(rep_df)} rows)")

# ─────────────────────────────────────────────
# TRENDS REPORT BUILDER
# ─────────────────────────────────────────────
# Trends report module — appended to stability_report_v2.py

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage

LOT_COLORS  = {"Lot1": "#2E75B6", "Lot2": "#ED7D31", "Validity": "#70AD47"}
LOT_MARKERS = {"Lot1": "o",       "Lot2": "s",        "Validity": "^"}

def tp_sort_key(tp):
    import re
    m = re.search(r'\d+', tp)
    return int(m.group()) if m else 0


def make_trend_chart(sample, lot, tp_labels, means, mean_spec, indiv_spec,
                     lot_type="Feasibility"):
    import matplotlib.lines as mlines

    LOT_COLORS = {"Lot1": "#2E75B6", "Lot2": "#C45911", "Validity": "#375623"}
    color = LOT_COLORS.get(lot, "#555555")

    valid_pairs = [(i, m) for i, m in enumerate(means) if m is not None]
    valid_x     = [p[0] for p in valid_pairs]
    valid_means = [p[1] for p in valid_pairs]

    fig, ax = plt.subplots(figsize=(9, 4.5))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#FAFAFA")

    # Shaded spec band and boundary lines
    if mean_spec:
        ax.axhspan(mean_spec["LSL"], mean_spec["USL"],
                   alpha=0.12, color="#70AD47", zorder=1)
        ax.axhline(mean_spec["LSL"], color="#375623", linewidth=1.8,
                   linestyle="-", zorder=2)
        ax.axhline(mean_spec["USL"], color="#375623", linewidth=1.8,
                   linestyle="-", zorder=2)
        ax.annotate(f'{mean_spec["USL"]}',
                    xy=(1, mean_spec["USL"]), xycoords=("axes fraction", "data"),
                    xytext=(4, 0), textcoords="offset points",
                    fontsize=8, color="#375623", va="center", fontweight="bold")
        ax.annotate(f'{mean_spec["LSL"]}',
                    xy=(1, mean_spec["LSL"]), xycoords=("axes fraction", "data"),
                    xytext=(4, 0), textcoords="offset points",
                    fontsize=8, color="#375623", va="center", fontweight="bold")

    # Individual rep spec lines
    if indiv_spec:
        ax.axhline(indiv_spec["LSL"], color="#C00000", linewidth=1.0,
                   linestyle="--", alpha=0.6, zorder=2)
        ax.axhline(indiv_spec["USL"], color="#C00000", linewidth=1.0,
                   linestyle="--", alpha=0.6, zorder=2)
        ax.annotate(f'{indiv_spec["USL"]}',
                    xy=(1, indiv_spec["USL"]), xycoords=("axes fraction", "data"),
                    xytext=(4, 0), textcoords="offset points",
                    fontsize=7.5, color="#C00000", va="center", alpha=0.7)
        ax.annotate(f'{indiv_spec["LSL"]}',
                    xy=(1, indiv_spec["LSL"]), xycoords=("axes fraction", "data"),
                    xytext=(4, 0), textcoords="offset points",
                    fontsize=7.5, color="#C00000", va="center", alpha=0.7)

    # Connection line
    if len(valid_x) > 1:
        ax.plot(valid_x, valid_means, color=color, linewidth=2,
                zorder=3, alpha=0.7)

    # Pass/fail coloured scatter points
    pass_x, pass_y, fail_x, fail_y = [], [], [], []
    for xi, m in zip(valid_x, valid_means):
        if mean_spec and mean_spec["LSL"] <= m <= mean_spec["USL"]:
            pass_x.append(xi); pass_y.append(m)
        else:
            fail_x.append(xi); fail_y.append(m)

    if pass_x:
        ax.scatter(pass_x, pass_y, color="#006100", s=80, zorder=6,
                   edgecolors="white", linewidths=1)
    if fail_x:
        ax.scatter(fail_x, fail_y, color="#9C0006", s=80, zorder=6,
                   marker="X", edgecolors="white", linewidths=1)

    # Value labels
    for xi, m in zip(valid_x, valid_means):
        in_spec   = mean_spec and mean_spec["LSL"] <= m <= mean_spec["USL"]
        txt_color = "#006100" if in_spec else "#9C0006"
        offset    = 8 if (not mean_spec or m < mean_spec["USL"] - 1) else -14
        ax.annotate(f"{m:.1f}", xy=(xi, m),
                    xytext=(0, offset), textcoords="offset points",
                    ha="center",
                    va="bottom" if offset > 0 else "top",
                    fontsize=8, color=txt_color, fontweight="bold", zorder=7)

    # Axes
    ax.set_xticks(list(range(len(tp_labels))))
    ax.set_xticklabels(tp_labels, fontsize=9, fontweight="bold")
    ax.set_xlabel("Timepoint", fontsize=10, labelpad=6)
    ax.set_ylabel("Mean Concentration (pmol/L)", fontsize=10)
    ax.set_title(f"{sample}  ·  {lot}", fontsize=12,
                 fontweight="bold", color=color, pad=10)

    # Y limits
    all_vals = valid_means[:]
    for s in [mean_spec, indiv_spec]:
        if s:
            all_vals += [s["LSL"], s["USL"]]
    if all_vals:
        span   = max(all_vals) - min(all_vals) or 10
        margin = span * 0.20
        ax.set_ylim(min(all_vals) - margin,
                    max(all_vals) + margin + span * 0.08)
    ax.set_xlim(-0.5, len(tp_labels) - 0.5)
    ax.set_position([0.10, 0.12, 0.82, 0.78])

    # Grid and spines
    ax.yaxis.grid(True, alpha=0.25, color="#AAAAAA")
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")

    # Legend
    handles = []
    if pass_x:
        handles.append(mlines.Line2D([], [], color="#006100", marker="o",
                                     linestyle="None", markersize=7, label="Pass"))
    if fail_x:
        handles.append(mlines.Line2D([], [], color="#9C0006", marker="X",
                                     linestyle="None", markersize=7, label="Fail"))
    if mean_spec:
        handles.append(mlines.Line2D([], [], color="#375623", linewidth=1.8,
                                     label=f"Mean Spec ({mean_spec['LSL']}–{mean_spec['USL']})"))
    if indiv_spec:
        handles.append(mlines.Line2D([], [], color="#C00000", linewidth=1,
                                     linestyle="--", alpha=0.7,
                                     label=f"Indiv Rep Spec ({indiv_spec['LSL']}–{indiv_spec['USL']})"))
    ax.legend(handles=handles, fontsize=8, loc="upper right",
              framealpha=0.85, edgecolor="#CCCCCC",
              bbox_to_anchor=(0.99, 0.99))

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()



def build_trends_report(study_folder, study_display, summary_df):
    import os, re, pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    out_path = os.path.join(
        study_folder,
        f"{study_display.replace(' ', '_')}_Trends_Report.xlsx"
    )

    active = summary_df[summary_df["AcceptedVer"] >= 0].copy()
    active["PassFail"] = active.apply(per_run_pass, axis=1)

    timepoints = sorted(active["Timepoint"].unique(), key=tp_sort_key)
    lots = [l for l in ["Lot1", "Lot2", "Validity"] if l in active["Lot"].unique()]

    wb = Workbook()

    # ── Overview sheet ──────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov.sheet_view.showGridLines = False
    ws_ov.column_dimensions["A"].width = 2

    ws_ov.merge_cells("B2:N3")
    t = ws_ov["B2"]
    t.value     = f"{study_display} — Stability Trends Report"
    t.font      = Font(bold=True, size=16, color="FFFFFF", name="Arial")
    t.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws_ov.row_dimensions[2].height = 28
    ws_ov.row_dimensions[3].height = 28

    meta = [("Study", study_display), ("Timepoints", ", ".join(timepoints)),
            ("Lots", ", ".join(lots)), ("Samples", ", ".join(SAMPLE_ORDER))]
    for i, (k, v) in enumerate(meta):
        r = 5 + i
        ws_ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws_ov.merge_cells(start_row=r, start_column=4, end_row=r, end_column=14)
        ws_ov.cell(row=r, column=2, value=k + ":").font = Font(bold=True, name="Arial", size=10)
        ws_ov.cell(row=r, column=4, value=v).font = Font(name="Arial", size=10)

    # Pass/fail summary table
    row = 11
    section_heading(ws_ov, row, 2, 3 + len(timepoints),
                    "Overall Pass/Fail Summary by Timepoint")
    row += 1
    apply_header(ws_ov, row, [2, 3], ["Sample", "Lot"])
    for ti, tp in enumerate(timepoints):
        apply_header(ws_ov, row, [4 + ti], [tp])
    row += 1

    for samp in SAMPLE_ORDER:
        for lot in lots:
            # Skip Validity rows for Panel samples — Panels are not
            # included in the validity lot, so these rows are always "—"
            if lot == "Validity" and samp.startswith("Panel"):
                continue
            vals = [samp, lot]
            for tp in timepoints:
                sub = active[(active["Sample"]==samp) &
                             (active["Lot"]==lot) &
                             (active["Timepoint"]==tp)]
                if sub.empty:
                    vals.append("—")
                else:
                    n_p = (sub["PassFail"]=="PASS").sum()
                    n_t = len(sub[sub["PassFail"].isin(["PASS","FAIL"])])
                    vals.append(f"{n_p}/{n_t}")
            for ci, v in enumerate(vals):
                pf = None
                if ci >= 2 and "/" in str(v):
                    np_, nt_ = v.split("/")
                    pf = ("PASS" if np_==nt_ and nt_!="0" else
                          "FAIL" if np_=="0" else None)
                wdc(ws_ov, row, ci + 2, v, pf=pf)
            row += 1

    ws_ov.column_dimensions["B"].width = 12
    ws_ov.column_dimensions["C"].width = 10
    for ti in range(len(timepoints)):
        ws_ov.column_dimensions[get_column_letter(4 + ti)].width = 10
    set_row_heights(ws_ov)

    # ── Per-sample sheets ───────────────────────────────────────────
    for samp in SAMPLE_ORDER:
        ws2 = wb.create_sheet(samp)
        ws2.sheet_view.showGridLines = False
        ws2.column_dimensions["A"].width = 2

        mean_spec  = MEAN_SPECS.get(samp, {})
        indiv_spec = INDIV_REP_SPECS.get(samp, {})
        spec_col   = 4 + len(timepoints)

        rw = 1
        section_heading(ws2, rw, 2, spec_col + 1,
                        f"{samp} — Concentration Trend Across Timepoints")
        rw += 1

        # Grand mean + CV table
        apply_header(ws2, rw, [2, 3], ["Lot", "Metric"])
        for ti, tp in enumerate(timepoints):
            apply_header(ws2, rw, [4 + ti], [tp])
        apply_header(ws2, rw, [spec_col, spec_col + 1],
                     ["Mean\nSpec Low", "Mean\nSpec High"])
        rw += 1

        chart_lots = []
        for lot in lots:
            # Validity lot is RefCon-only — skip for Panel samples
            if lot == "Validity" and samp.startswith("Panel"):
                continue
            lot_means = []
            # Grand mean row
            wdc(ws2, rw, 2, lot, bold=True)
            wdc(ws2, rw, 3, "Grand Mean")
            for ti, tp in enumerate(timepoints):
                sub = active[(active["Sample"]==samp) &
                             (active["Lot"]==lot) &
                             (active["Timepoint"]==tp)]
                if sub.empty:
                    lot_means.append(None)
                    wdc(ws2, rw, 4 + ti, None)
                else:
                    gm = round_conc(sub["MeanDose"].dropna().mean())
                    lot_means.append(gm)
                    ok = mean_spec and mean_spec["LSL"] <= gm <= mean_spec["USL"]
                    wdc(ws2, rw, 4 + ti, gm,
                        pf="PASS" if ok else "FAIL", fmt=conc_fmt())
            wdc(ws2, rw, spec_col,     mean_spec.get("LSL"), fmt=conc_fmt())
            wdc(ws2, rw, spec_col + 1, mean_spec.get("USL"), fmt=conc_fmt())
            rw += 1

            # %CV row
            wdc(ws2, rw, 2, lot, bold=True)
            wdc(ws2, rw, 3, "%CV")
            for ti, tp in enumerate(timepoints):
                sub = active[(active["Sample"]==samp) &
                             (active["Lot"]==lot) &
                             (active["Timepoint"]==tp)]
                if sub.empty:
                    wdc(ws2, rw, 4 + ti, None)
                else:
                    vals_cv = sub["MeanDose"].dropna()
                    cv = round(vals_cv.std() / vals_cv.mean() * 100, 2) \
                         if len(vals_cv) >= 2 else None
                    ok = cv is not None and cv <= CV_LIMIT
                    wdc(ws2, rw, 4 + ti, cv,
                        pf="PASS" if ok else ("FAIL" if cv is not None else None),
                        fmt="0.00")
            wdc(ws2, rw, spec_col,     f"CV ≤ {CV_LIMIT}%")
            wdc(ws2, rw, spec_col + 1, "—")
            rw += 1

            chart_lots.append((lot, lot_means))

        rw += 1

        # Charts — one per lot stacked vertically
        for lot, lot_means in chart_lots:
            png = make_trend_chart(samp, lot, timepoints, lot_means,
                                   mean_spec or None, indiv_spec or None)
            img = XLImage(BytesIO(png))
            img.width  = 540
            img.height = 270
            ws2.add_image(img, f"B{rw}")
            rw += 18

        # Column widths
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 12
        for ti in range(len(timepoints)):
            ws2.column_dimensions[get_column_letter(4 + ti)].width = 12
        ws2.column_dimensions[get_column_letter(spec_col)].width = 12
        ws2.column_dimensions[get_column_letter(spec_col + 1)].width = 12
        set_row_heights(ws2)

    wb.save(out_path)
    print(f"  ✅ Trends report saved: {out_path}")
    return out_path


def run_study(study_folder):
    """
    Process an entire study folder.
    Generates one per-TP report per timepoint found.
    Trends report will be added in next phase.
    """
    print(f"\nProcessing study: {study_folder}")
    study_display, summary_df, reps_df = load_study(study_folder)

    if summary_df.empty:
        print("❌ No data loaded.")
        return

    timepoints = sorted(summary_df["Timepoint"].unique())
    print(f"\nTimepoints found: {timepoints}")

    for tp in timepoints:
        out_name = f"{study_display.replace(' ', '_')}_{tp}_Report.xlsx"
        out_path = os.path.join(study_folder, out_name)
        print(f"\nBuilding {tp} report...")
        build_tp_report(summary_df, reps_df, study_display, tp, out_path)


    # Build trends report across all timepoints
    print(f"\nBuilding trends report...")
    build_trends_report(study_folder, study_display, summary_df)

    # Export Tableau CSVs
    print(f"\nExporting Tableau data...")
    export_tableau_csv(study_folder, study_display, summary_df, reps_df)

    print(f"\n✅ All reports complete.")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        run_study(sys.argv[1])
    else:
        print("Usage: python stability_report_v2.py <study_folder>")
        print("Example: python stability_report_v2.py HE4_Stab11")
